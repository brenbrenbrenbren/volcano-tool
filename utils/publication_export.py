"""
Publication-ready figure export with multi-panel layouts.

Generate publication-quality figures with multiple panels.
"""

import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
from typing import Dict, List, Optional
from datetime import datetime


class PublicationExporter:
    """Export publication-ready multi-panel figures."""

    def __init__(self):
        """Initialize exporter."""
        pass

    def create_multi_panel_figure(
        self,
        panels: List[Dict],
        layout: str = "2x2",
        title: str = None,
        width: int = 1200,
        height: int = 1000
    ) -> go.Figure:
        """
        Create a multi-panel figure for publication.

        Args:
            panels: List of panel dictionaries with 'figure', 'title', 'label' keys
            layout: Layout pattern (e.g., "2x2", "1x3", "3x1")
            title: Overall figure title
            width: Figure width in pixels
            height: Figure height in pixels

        Returns:
            Combined Plotly figure
        """
        # Parse layout
        rows, cols = map(int, layout.split('x'))

        # Create subplot figure
        subplot_titles = [p.get('title', '') for p in panels[:rows*cols]]

        fig = make_subplots(
            rows=rows,
            cols=cols,
            subplot_titles=subplot_titles,
            vertical_spacing=0.12,
            horizontal_spacing=0.1,
            specs=[[{"type": "xy"} for _ in range(cols)] for _ in range(rows)]
        )

        # Add each panel
        for idx, panel in enumerate(panels[:rows*cols]):
            row = idx // cols + 1
            col = idx % cols + 1

            panel_fig = panel['figure']

            # Add traces from panel to subplot
            for trace in panel_fig.data:
                fig.add_trace(trace, row=row, col=col)

            # Add panel label (A, B, C, etc.)
            if 'label' in panel:
                fig.add_annotation(
                    text=f"<b>{panel['label']}</b>",
                    xref=f"x{idx+1} domain",
                    yref=f"y{idx+1} domain",
                    x=0,
                    y=1.1,
                    showarrow=False,
                    font=dict(size=18, family="Arial Black"),
                    xanchor='left'
                )

        # Update layout
        fig.update_layout(
            title=dict(text=title, font=dict(size=20)) if title else None,
            width=width,
            height=height,
            showlegend=False,
            font=dict(family="Arial", size=12),
            paper_bgcolor='white',
            plot_bgcolor='white'
        )

        return fig

    def generate_figure_legend(
        self,
        panels: List[Dict],
        figure_number: int = 1,
        include_methods: bool = True
    ) -> str:
        """
        Generate figure legend text for publication.

        Args:
            panels: List of panel dictionaries
            figure_number: Figure number
            include_methods: Include methods summary

        Returns:
            Formatted legend text
        """
        legend = f"Figure {figure_number}. "

        # Overall figure description
        if len(panels) > 0 and 'description' in panels[0]:
            legend += panels[0]['description'] + "\n"

        # Panel descriptions
        for i, panel in enumerate(panels):
            label = panel.get('label', chr(65 + i))  # A, B, C, etc.
            desc = panel.get('panel_description', f"Panel {label}")
            legend += f"\n({label}) {desc}"

        # Methods summary
        if include_methods:
            legend += "\n\nDifferential expression analysis performed using "
            legend += "DESeq2-like statistics. Genes with |log2FC| ≥ 1.0 and "
            legend += "adjusted p-value ≤ 0.05 considered significant. "
            legend += f"Generated on {datetime.now().strftime('%Y-%m-%d')}."

        return legend

    def create_figure_package(
        self,
        panels: List[Dict],
        figure_number: int = 1,
        layout: str = "2x2",
        title: str = None
    ) -> Dict:
        """
        Create complete figure package with figure and legend.

        Args:
            panels: List of panel dictionaries
            figure_number: Figure number
            layout: Layout pattern
            title: Overall figure title

        Returns:
            Dictionary with 'figure' and 'legend' keys
        """
        # Create multi-panel figure
        fig = self.create_multi_panel_figure(
            panels=panels,
            layout=layout,
            title=title
        )

        # Generate legend
        legend = self.generate_figure_legend(
            panels=panels,
            figure_number=figure_number
        )

        return {
            'figure': fig,
            'legend': legend,
            'panels': len(panels),
            'layout': layout
        }

    def export_supplementary_table(
        self,
        data: pd.DataFrame,
        filename: str = "supplementary_table.xlsx",
        sheet_name: str = "Data",
        legend: str = None
    ) -> bytes:
        """
        Export data as publication-ready supplementary table.

        Args:
            data: DataFrame to export
            filename: Output filename
            sheet_name: Excel sheet name
            legend: Table legend/description

        Returns:
            Excel file as bytes
        """
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        # Write to Excel in memory
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data
            data.to_excel(writer, sheet_name=sheet_name, index=False)

            # Style the workbook
            workbook = writer.book
            worksheet = workbook[sheet_name]

            # Header styling
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add legend if provided
            if legend:
                legend_sheet = workbook.create_sheet("Legend")
                legend_sheet['A1'] = "Table Legend"
                legend_sheet['A1'].font = Font(bold=True, size=14)
                legend_sheet['A3'] = legend
                legend_sheet['A3'].alignment = Alignment(wrap_text=True, vertical='top')
                legend_sheet.column_dimensions['A'].width = 100

        output.seek(0)
        return output.getvalue()
